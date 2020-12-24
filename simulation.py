import sys
import os

import json
import threading
import time
import numpy as np

from pymongo import MongoClient
import datetime
from openpyxl import Workbook

import upbitAPI

from matplotlib import pyplot as plt
from collections import deque

_ACC_VOL = "candle_acc_trade_volume"
_TIME_UTC = "candle_date_time_utc"

def printData(prefix, data):
    print(prefix, end=' ')
    print(f"vol:{data[_ACC_VOL]:.2f}", end='\t')
    print(f"open:{data['opening_price']:.2f}", end='\t')
    print(f"low:{data['low_price']:.2f}", end='\t')
    print(f"high:{data['high_price']:.2f}", end='\t')    
    print(f"price:{data['trade_price']:.2f}")

if __name__ == '__main__':

    with open('dbinfo.json', 'r') as f:
        dbinfo = f.read()
    dbinfo = json.loads(dbinfo)

    with open('key.json', 'r') as f:
        key = f.read()
    key = json.loads(key)

    db_client = MongoClient(dbinfo['addr'])
    db_set = db_client[dbinfo['name']]
    db_collector = db_set[dbinfo['collector']]

    upbit_api = upbitAPI.upbitAPI(key)

    coin_list = upbit_api.req_coinlist()

    simul_info = db_collector.find({ "type":"simulinfo"})
    simul_info = simul_info[0]

    # for i in range(len(simul_info['coinlist'])):
        # print(simul_info['coinlist'][i], simul_info['coinname'][i])

    test_coin = 'KRW-SBD'
    avg_day = 5
    th_benefit_ratio = 1.10


    print(test_coin)

    min_data_set = list(db_collector.find({"market":test_coin, "unit": 5}).sort(_TIME_UTC))
    day_data_set = list(db_collector.find({"market":test_coin, "unit": 'day'}).sort(_TIME_UTC))

    price_set = []
    volume_ratio_set = []
    volume_set = deque()
    price_set = deque()
    print_next = 0

    write_wb = Workbook()
    write_ws_rate = write_wb.create_sheet(test_coin + '_rate')
    write_ws_dayvolume = write_wb.create_sheet(test_coin + '_dayv')

    write_ws_rate.cell(row=1, column=1).value = 'time'
    write_ws_rate.cell(row=1, column=2).value = 'volume'
    write_ws_rate.cell(row=1, column=3).value = str(avg_day) + 'volume'
    write_ws_rate.cell(row=1, column=4).value = 'open'
    write_ws_rate.cell(row=1, column=5).value = 'low'
    write_ws_rate.cell(row=1, column=6).value = 'high'
    write_ws_rate.cell(row=1, column=7).value = 'price'
    write_ws_rate.cell(row=1, column=8).value = str(avg_day) + 'price'
    write_ws_rate.cell(row=1, column=9).value = 'rate'
    write_ws_rate.cell(row=1, column=10).value = 'n rate'
    write_ws_rate.cell(row=1, column=11).value = 'v rate'
    write_ws_rate.cell(row=1, column=12).value = 'day volume'
    write_ws_rate.cell(row=1, column=13).value = 'day price'

    write_ws_dayvolume.cell(row=1, column=1).value = 'time'
    write_ws_dayvolume.cell(row=1, column=2).value = 'volume'
    write_ws_dayvolume.cell(row=1, column=3).value = str(avg_day) + 'volume'
    write_ws_dayvolume.cell(row=1, column=4).value = 'open'
    write_ws_dayvolume.cell(row=1, column=5).value = 'low'
    write_ws_dayvolume.cell(row=1, column=6).value = 'high'
    write_ws_dayvolume.cell(row=1, column=7).value = 'price'
    write_ws_dayvolume.cell(row=1, column=8).value = str(avg_day) + 'price'
    write_ws_dayvolume.cell(row=1, column=9).value = 'rate'
    write_ws_dayvolume.cell(row=1, column=10).value = 'n rate'
    write_ws_dayvolume.cell(row=1, column=11).value = 'v rate'
    write_ws_dayvolume.cell(row=1, column=12).value = 'day volume'
    write_ws_dayvolume.cell(row=1, column=13).value = 'day price'
    
    ws_rate_row_cnt = 2
    ws_dayv_row_cnt = 2
    for i in range(1,len(min_data_set)-1):
        min_data = min_data_set[i]
        min_data_prev = min_data_set[i-1]
        min_data_next = min_data_set[i+1]
        min_date_str = min_data[_TIME_UTC]
        date = datetime.datetime.strptime(min_date_str, '%Y-%m-%dT%H:%M:%S')
        date = date - datetime.timedelta(days=1)
        date_str = date.strftime('%Y-%m-%dT%H:%M:%S')
        date_str = date_str[0:11] + '00:00:00'
        
        finded = 0
        idx = 0
        day_data = dict()
        while finded == 0 and idx < len(day_data_set):
            if day_data_set[idx][_TIME_UTC] == date_str:
                finded = 1
                day_data = day_data_set[idx]
            idx += 1

        if len(day_data) <= 0:
            continue

        price = min_data['trade_price']
        high_price = min_data['high_price']
        low_price = min_data['low_price']
        open_price = min_data['opening_price']
        volume = min_data[_ACC_VOL]

        if len(volume_set) < avg_day:
            volume_set.append(volume)
            price_set.append(price)
            continue
        else:
            volume_set.rotate(1)
            volume_set[0] = volume
            price_set.rotate(1)
            price_set[0] = price

        avr_volume = np.mean(volume_set)
        avr_price = np.mean(price_set)
        day_volume = day_data[_ACC_VOL]
        day_price = day_data['trade_price']
        high_price_ratio = high_price/price

        benefit_ratio = min_data['trade_price']/min_data['opening_price']
        benefit_ratio_est = min_data_next['trade_price']/min_data_next['opening_price']
        # if benefit_ratio > th_benefit_ratio:
        #     print('-'*5, min_date_str, '-'*5)
        #     print(f'benefit ratio:{benefit_ratio:.2f} next:{benefit_ratio_est:.2f}')
        #     printData("PREV", min_data_prev)
        #     printData("CUR ", min_data)
        #     printData("NEXT", min_data_next)
        #     print(f"DAY  vol:{day_data[_ACC_VOL]:.2f}\trate:{day_data['change_rate']}\tprice:{day_data['trade_price']}")
        #     print(f"STAT avr volume:{avr_volume:.2f}\tavr price:{avr_price:.2f}")      


        volume_ratio = volume/avr_volume
        volume_day_ratio = volume/day_volume
        # if volume_ratio > 2:
        if benefit_ratio_est > th_benefit_ratio:
            write_ws_rate.cell(row=ws_rate_row_cnt, column=1).value = min_date_str
            write_ws_rate.cell(row=ws_rate_row_cnt, column=2).value = volume
            write_ws_rate.cell(row=ws_rate_row_cnt, column=3).value = avr_volume
            write_ws_rate.cell(row=ws_rate_row_cnt, column=4).value = open_price
            write_ws_rate.cell(row=ws_rate_row_cnt, column=5).value = low_price
            write_ws_rate.cell(row=ws_rate_row_cnt, column=6).value = high_price
            write_ws_rate.cell(row=ws_rate_row_cnt, column=7).value = price
            write_ws_rate.cell(row=ws_rate_row_cnt, column=8).value = avr_price
            write_ws_rate.cell(row=ws_rate_row_cnt, column=9).value = benefit_ratio
            write_ws_rate.cell(row=ws_rate_row_cnt, column=10).value = benefit_ratio_est
            write_ws_rate.cell(row=ws_rate_row_cnt, column=11).value = volume_ratio
            write_ws_rate.cell(row=ws_rate_row_cnt, column=12).value = day_volume
            write_ws_rate.cell(row=ws_rate_row_cnt, column=13).value = day_price

            ws_rate_row_cnt += 1

        if volume_day_ratio > 1:
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=1).value = min_date_str
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=2).value = volume
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=3).value = avr_volume
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=4).value = open_price
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=5).value = low_price
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=6).value = high_price
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=7).value = price
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=8).value = avr_price
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=9).value = benefit_ratio
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=10).value = benefit_ratio_est
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=11).value = volume_ratio
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=12).value = day_volume
            write_ws_dayvolume.cell(row=ws_dayv_row_cnt, column=13).value = day_price

            ws_dayv_row_cnt += 1

        # if print_next == 1:
        #     # print(f'Next avr_v:{avr_volume} day_v:{day_volume} p:{price} avr_p:{avr_price} high_p:{high_price} day_p:{day_price}')
        #     benefit_ratio = price/open_price
        #     if benefit_ratio > 1.10:
        #         print("-"*20, date_str)
        #         print(f'benefit ratio{benefit_ratio}')
        #         print(f'avr_v:{avr_volume} day_v:{day_volume} p:{price} avr_p:{avr_price} high_p:{high_price} day_p:{day_price}')
        #         print(prv_data)
        #     print_next = 0

        # if volume > day_volume*ratio_th_volume:
        #     # print("-"*20)
        #     # print(date_str)
        #     # print(f'avr_v:{avr_volume} day_v:{day_volume} p:{price} avr_p:{avr_price} high_p:{high_price} day_p:{day_price}')
        #     # print(prv_data)
        #     print_next = 1

        # benefit_ratio = price/open_price
        # if benefit_ratio > 1.10:
        #     print("-"*20, date_str)
        #     print(f'benefit ratio:{benefit_ratio:.2f}')
        #     print(f'day volume:{day_volume} day price:{day_price}')
        #     print(f'volume:{volume:.2f} / {prv_v:.2f}')
        #     print(f'avr volume:{avr_volume:.2f} / {prv_avr_v:.2f}')
        #     print(f'price:{price} / {prv_p}')
        #     print(f'avr price:{avr_price:.2f} / {prv_avr_p:.2f}')
        #     print(f'high price:{high_price} / {prv_high_p}')
        #     print(f'low price:{low_price} / {prv_low_p}')

        price_set.append(min_data['trade_price'])
        volume_ratio_set.append(volume_ratio)

    avr_volume_ratio = np.mean(volume_ratio_set)
    std_volume_ratio = np.std(volume_ratio_set)
    var_volume_ratio = np.var(volume_ratio_set)

    print(avr_volume_ratio, std_volume_ratio, var_volume_ratio)

    write_wb.remove(write_wb['Sheet'])
    excel_name = test_coin + '_avg' + str(avg_day) + '.xlsx'
    write_wb.save(excel_name)
    write_wb.close()

    # plt.plot(volume_ratio_set)
    # plt.show()

    sys.exit()