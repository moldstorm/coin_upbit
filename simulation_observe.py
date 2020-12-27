import sys
import os

import json
import threading
import time
import numpy as np

from pymongo import MongoClient
import datetime
from openpyxl import Workbook

import upbitAPI

from matplotlib import pyplot as plt
from collections import deque

from tkinter import *
from tkinter import messagebox

_ACC_VOL = "candle_acc_trade_volume"
_TIME_UTC = "candle_date_time_utc"

def printData(prefix, data):
    print(prefix, end=' ')
    print(f"vol:{data[_ACC_VOL]:.2f}", end='\t')
    print(f"open:{data['opening_price']:.2f}", end='\t')
    print(f"low:{data['low_price']:.2f}", end='\t')
    print(f"high:{data['high_price']:.2f}", end='\t')    
    print(f"price:{data['trade_price']:.2f}")

if __name__ == '__main__':

    with open('dbinfo.json', 'r') as f:
        dbinfo = f.read()
    dbinfo = json.loads(dbinfo)

    with open('key.json', 'r') as f:
        key = f.read()
    key = json.loads(key)

    db_client = MongoClient(dbinfo['addr'])
    db_set = db_client[dbinfo['name']]
    db_collector = db_set[dbinfo['collector']]

    upbit_api = upbitAPI.upbitAPI(key)

    coin_list = upbit_api.req_coinlist()

    simul_info = db_collector.find({ "type":"simulinfo"})
    simul_info = simul_info[0]

    
    avg_cnt = 5
    gain_val = 1.05
    th_change_rate_low_minute = 1.20
    th_loss_benefit_ratio = 0.95

    th_change_rate_low_day = 1.20

    multiple_volume_min = 10

    # th_rate_low = 1.05
    # th_rate_high = 1.1
    # th_c_rate_low = 1.05
    # th_c_rate_high = 1.1
    # th_v_rate_high = 150
    th_nc_rate_high = 1.2

    th_v_rate_low = 10
    th_day_v_rate_high = 1.0

    th_price = 110
    

    # for i in range(len(simul_info['coinlist'])):
        # print(simul_info['coinlist'][i], simul_info['coinname'][i])

    write_wb = Workbook()

    write_ws_comb = write_wb.create_sheet('gain '+str(th_change_rate_low_minute))

    write_ws_comb.cell(row=1, column=1).value = 'time'
    write_ws_comb.cell(row=1, column=2).value = 'market'
    write_ws_comb.cell(row=1, column=3).value = 'case'
    write_ws_comb.cell(row=1, column=4).value = 'volume'
    write_ws_comb.cell(row=1, column=5).value = str(avg_cnt) + 'volume'
    write_ws_comb.cell(row=1, column=6).value = 'open'
    write_ws_comb.cell(row=1, column=7).value = 'low'
    write_ws_comb.cell(row=1, column=8).value = 'high'
    write_ws_comb.cell(row=1, column=9).value = 'price'
    write_ws_comb.cell(row=1, column=10).value = 'day volume'
    write_ws_comb.cell(row=1, column=11).value = 'day open'
    write_ws_comb.cell(row=1, column=12).value = 'day high'
    write_ws_comb.cell(row=1, column=13).value = 'day price'
    write_ws_comb.cell(row=1, column=14).value = 'volume rate' 
    write_ws_comb.cell(row=1, column=15).value = 'end rate' 
    write_ws_comb.cell(row=1, column=16).value = 'start rate' 
    
    ws_rate_row_cnt = 2
    ws_dayv_row_cnt = 2
    ws_comb_row_cnt = 2    

    # coin_list = []
    # test_coin = 'KRW-SBD'
    # coin_list.append(dict(market=test_coin))

    # file_ptr = open("volume_check.log", 'w')

    for i in range(len(coin_list)):
        test_coin = coin_list[i]['market']
        if test_coin.find('KRW') < 0:
            continue

        print(f'[{i}/{len(coin_list)}] {test_coin}')
        # file_ptr.write(test_coin + '\n')

        minute_data_set = list(db_collector.find({"market":test_coin, "unit": 5}).sort(_TIME_UTC))
        day_data_set = list(db_collector.find({"market":test_coin, "unit": 'day'}).sort(_TIME_UTC))

        for i in range(len(day_data_set)):
            day_data = day_data_set[i]
            day_change_rate = day_data['trade_price']/day_data['opening_price']

            if day_change_rate >= th_change_rate_low_day:
                day_date_str = day_data[_TIME_UTC]
                date = datetime.datetime.strptime(day_date_str, '%Y-%m-%dT%H:%M:%S')
                day_str = date.strftime('%Y-%m-%d')

                day_high_price = day_data['high_price']

                print(f'[{day_str}] rate:{day_change_rate:.2f}% open:{day_data["opening_price"]} cur:{day_data["trade_price"]} high:{day_high_price}')

                minute_start_day_idx = -1
                j = 0
                while minute_start_day_idx == -1 and j < len(minute_data_set):
                    minute_data = minute_data_set[j]
                    minute_date_str = minute_data[_TIME_UTC]
                    date = datetime.datetime.strptime(minute_date_str, '%Y-%m-%dT%H:%M:%S')
                    minute_day_str = date.strftime('%Y-%m-%d')
                    if minute_day_str == day_str:
                        minute_start_day_idx = j
                    j+= 1

                if minute_start_day_idx is not -1:
                    minute_high_idx = -1            
                    minute_high_price = -1

                    j = minute_start_day_idx
                    while minute_high_idx == -1 and j < len(minute_data_set):
                        minute_data = minute_data_set[j]
                        if day_high_price == minute_data['high_price']:
                            minute_high_idx = j
                            minute_high_price = minute_data['high_price']
                            minute_high_trade_price = minute_data['trade_price']
                        j -= 1

                    if not minute_high_idx == -1:

                        volume_set = deque()
                        price_set = deque()

                        write_on = 0
                        start_price = day_data['opening_price']

                        for j in range(minute_high_idx - 48, minute_high_idx+1):
                            minute_data = minute_data_set[j]
                            minute_date_str = minute_data[_TIME_UTC]
                            date = datetime.datetime.strptime(minute_date_str, '%Y-%m-%dT%H:%M:%S')
                            time_str = date.strftime('%H:%M:%S')

                            price = minute_data['trade_price']
                            high_price = minute_data['high_price']
                            low_price = minute_data['low_price']
                            open_price = minute_data['opening_price']
                            volume = minute_data[_ACC_VOL]

                            if len(volume_set) < avg_cnt + 1:
                                volume_set.append(volume)
                                price_set.append(price)
                                continue
                            else:
                                volume_set.rotate(1)
                                volume_set[0] = volume
                                price_set.rotate(1)
                                price_set[0] = price
                                volume_set_past = list(volume_set)
                                volume_set_past = volume_set_past[1:avg_cnt+1]
                                price_set_past = list(price_set)
                                price_set_past = price_set_past[1:avg_cnt+1]

                            avr_volume = np.mean(volume_set_past)

                            ratio_volume = volume/avr_volume

                            next_price = minute_data_set[j+1]['trade_price']

                            if th_v_rate_low <= ratio_volume and (minute_high_price/price) > (th_change_rate_low_minute - 0.05):
                                write_on = 1
                                start_price = price

                            if write_on == 1:
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=1).value = minute_date_str
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=2).value = test_coin
                                # write_ws_comb.cell(row=ws_comb_row_cnt, column=3).value = 'case'
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=4).value = 'volume'
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=5).value = avr_volume
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=6).value = open_price
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=7).value = low_price
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=8).value = high_price
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=9).value = price
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=10).value = day_data[_ACC_VOL]
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=11).value = day_data['opening_price']
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=12).value = day_data['high_price']
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=13).value = day_data['trade_price']
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=14).value = ratio_volume
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=15).value = minute_high_trade_price/price
                                write_ws_comb.cell(row=ws_comb_row_cnt, column=16).value = price/start_price

                                ws_comb_row_cnt += 1                    
                                print(f'[{time_str}] ratio:{ratio_volume:.4f} cur:{price} next:{next_price} high:{high_price} rate:{minute_high_trade_price/price:.4f}')

                            # print_f = f'[{time_str}] ratio:{ratio_volume:.4f} cur:{price} next:{next_price} high:{high_price}\n'
                            # file_ptr.write(print_f)
                    # minute_ratio = 0
                    # j = minute_high_idx
                    # while minute_ratio <= day_change_rate:
                    #     minute_ratio = minute_high_price/minute_data_set[j]['trade_price']
                    #     j -= 1

                    # print(f"ratio:{minute_ratio} idx:{j}/{minute_high_idx}")
                    # volume_set = deque()
                    # price_set = deque()
                    # print_next = 0

                    # over_day = -1
                    # j = minute_start_day_idx
                    # while over_day == -1:
                    #     minute_data = minute_data_set[j]
                    #     minute_date_str = minute_data[_TIME_UTC]
                    #     date = datetime.datetime.strptime(minute_date_str, '%Y-%m-%dT%H:%M:%S')
                    #     minute_day_str = date.strftime('%Y-%m-%d')
                    #     if not (minute_day_str == day_str):
                    #         over_day = 1
                    #         continue

                    #     if j+1 == len(minute_data_set):
                    #         over_day = 1
                    #         continue

                    #     j += 1

                    #     price = minute_data['trade_price']
                    #     high_price = minute_data['high_price']
                    #     low_price = minute_data['low_price']
                    #     open_price = minute_data['opening_price']
                    #     volume = minute_data[_ACC_VOL]

                    #     if len(volume_set) < avg_cnt + 1:
                    #         volume_set.append(volume)
                    #         price_set.append(price)
                    #         continue
                    #     else:
                    #         volume_set.rotate(1)
                    #         volume_set[0] = volume
                    #         price_set.rotate(1)
                    #         price_set[0] = price
                    #         volume_set_past = list(volume_set)
                    #         volume_set_past = volume_set_past[1:avg_cnt+1]
                    #         price_set_past = list(price_set)
                    #         price_set_past = price_set_past[1:avg_cnt+1]

                    #     avr_volume = np.mean(volume_set_past)
                    #     avr_price = np.mean(price_set_past)

                    #     if volume >= multiple_volume_min*avr_volume:
                    #         next_price = minute_data_set[j+1]['trade_price']
                    #         print(f'cur:{price} next:{next_price}')
                        
        # price_set = []
        # volume_ratio_set = []
        # volume_set = deque()
        # price_set = deque()
        # print_next = 0

        # for i in range(1,len(minute_data_set)-1):
        #     minute_data = minute_data_set[i]
        #     minute_data_prev = minute_data_set[i-1]
        #     minute_data_next = minute_data_set[i+1]
        #     minute_date_str = minute_data[_TIME_UTC]
        #     date = datetime.datetime.strptime(minute_date_str, '%Y-%m-%dT%H:%M:%S')
        #     date = date - datetime.timedelta(days=1)
        #     date_str = date.strftime('%Y-%m-%dT%H:%M:%S')
        #     date_str = date_str[0:11] + '00:00:00'
            
        #     finded = 0
        #     idx = 0
        #     day_data = dict()
        #     while finded == 0 and idx < len(day_data_set):
        #         if day_data_set[idx][_TIME_UTC] == date_str:
        #             finded = 1
        #             day_data = day_data_set[idx]
        #         idx += 1

        #     if len(day_data) <= 0:
        #         continue

        #     price = minute_data['trade_price']
        #     high_price = minute_data['high_price']
        #     low_price = minute_data['low_price']
        #     open_price = minute_data['opening_price']
        #     volume = minute_data[_ACC_VOL]

        #     if len(volume_set) < avg_cnt + 1:
        #         volume_set.append(volume)
        #         price_set.append(price)
        #         continue
        #     else:
        #         volume_set.rotate(1)
        #         volume_set[0] = volume
        #         price_set.rotate(1)
        #         price_set[0] = price
        #         volume_set_past = list(volume_set)
        #         volume_set_past = volume_set_past[1:avg_cnt+1]
        #         price_set_past = list(price_set)
        #         price_set_past = price_set_past[1:avg_cnt+1]

        #     avr_volume = np.mean(volume_set_past)
        #     avr_price = np.mean(price_set_past)
        #     day_volume = day_data[_ACC_VOL]
        #     day_price = day_data['trade_price']
        #     high_price_ratio = high_price/price

        #     benefit_ratio = minute_data['trade_price']/minute_data['opening_price']
        #     benefit_ratio_est = minute_data_next['trade_price']/minute_data_next['opening_price']

        #     volume_ratio = volume/avr_volume
        #     volume_day_ratio = volume/day_volume
        #     change_rate = high_price/low_price

        #     if price == low_price:
        #         norm_change_rate = high_price/price
        #     else:
        #         norm_change_rate = (high_price-low_price)/(price-low_price)

        #     # if (volume_day_ratio > 1) and (benefit_ratio_est > th_benefit_ratio):
        #     # if benefit_ratio_est > th_benefit_ratio and th_price <= price:
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=1).value = minute_date_str
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=2).value = test_coin
        #     #     if (benefit_ratio_est > th_benefit_ratio):
        #     #         write_ws_comb.cell(row=ws_comb_row_cnt, column=3).value = 'gain'
        #     #     else:
        #     #         write_ws_comb.cell(row=ws_comb_row_cnt, column=3).value = 'loss'
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=4).value = volume
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=5).value = avr_volume
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=6).value = open_price
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=7).value = low_price
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=8).value = high_price
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=9).value = price
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=10).value = avr_price
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=11).value = day_volume
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=12).value = change_rate
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=13).value = norm_change_rate
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=14).value = benefit_ratio
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=15).value = volume_ratio                
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=16).value = volume_day_ratio
        #     #     write_ws_comb.cell(row=ws_comb_row_cnt, column=17).value = benefit_ratio_est

        #     #     ws_comb_row_cnt += 1


        #     if th_price <= price and \
        #         th_v_rate_low <= volume_ratio and th_day_v_rate_high >= volume_day_ratio and \
        #         th_nc_rate_high >= norm_change_rate:
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=1).value = minute_date_str
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=2).value = test_coin
        #         if (benefit_ratio_est >= gain_val):
        #             write_ws_comb.cell(row=ws_comb_row_cnt, column=3).value = 'gain'
        #         else:
        #             write_ws_comb.cell(row=ws_comb_row_cnt, column=3).value = 'loss'
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=4).value = volume
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=5).value = avr_volume
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=6).value = open_price
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=7).value = low_price
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=8).value = high_price
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=9).value = price
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=10).value = avr_price
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=11).value = day_volume
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=12).value = change_rate
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=13).value = norm_change_rate
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=14).value = benefit_ratio
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=15).value = volume_ratio
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=16).value = volume_day_ratio
        #         write_ws_comb.cell(row=ws_comb_row_cnt, column=17).value = benefit_ratio_est

        #         ws_comb_row_cnt += 1



    print('write result')

    write_wb.remove(write_wb['Sheet'])
    excel_name = 'collect' + '.xlsx'
    try:
        write_wb.save(excel_name)
        write_wb.close()
    except:
        print('error write excel')

    # plt.plot(volume_ratio_set)
    # plt.show()

    # file_ptr.close()

    messagebox.showinfo("Done", "Processing Done")

    sys.exit()